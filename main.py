# Dependencias Kivy:
import kivy
kivy.require('1.9.0')
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
## from kivy.uix.widget import Widget
from kivy.uix.popup import Popup
from kivy.properties import StringProperty, ObjectProperty
from kivy.config import Config
from kivy.uix.screenmanager import Screen, ScreenManager, FadeTransition
from plyer import filechooser # para ubuntu

# Dependencias bases de datos:
from peewee import  SqliteDatabase, Model, DateField, FloatField, CharField, CompositeKey
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# Otras dependencias
import time
import os
from platform import system
from tkinter import filedialog # tkinter para windows
import configparser
import threading
import re

# valores defoult
RUTA_DIR = os.getcwd()
FECHA_SIS = time.strftime("%d/%m/%Y", time.localtime(time.time()))
NOM_CFG = 'segpeso.cfg'
NOM_DEFOULT = {
    "xlsx": "registro_medidas.xlsx",
    "xlsx_pr": "prueba.xlsx",
    "bd_sqlite":"registro.db"}

# Controlar ventanas emergentes
global nom_inval_pop, sin_camp_pop 
nom_inval_pop, sin_camp_pop = None, None

# Verificar configuración previa
def buscar_cfg():
    global _config
    if NOM_CFG in os.listdir():
        _config = True
    else:
        _config = False

buscar_cfg()


def hora() -> str:
    '''Para generar claves primarias.'''
    hora = time.strftime("%d/%m/%Y-%H:%M:%S", time.localtime(time.time()))
    return hora

class PrimEjec(Screen):
    def __init__(self, **kw):
        global _config
        super().__init__(**kw)
        self.config = Confg()

    def confg_defoult(self):
        global inicio
        print("\nSeleccionada configuración por defecto. Iniciando APP.")
        self.config.cfg_defoult()
        buscar_cfg()
        time.sleep(1)
        inicio.add_widget(PesoApp(name = "app"))
        inicio.current = "app"


class CargarArch(Screen):
    '''Pantalla para buscar archivos .xlsx previos.'''
    ruta_arch = StringProperty()
    
    def __init__(self, **kw):
        super().__init__(**kw)
        self.ruta_arch = "\n    Buscar..."
        self.conf = Confg()
        self.habil_confirm = False
    
    def buscar_arch(self):
        try:
            self.ruta_arch = filedialog.askopenfilename()
            if self.ruta_arch:
                ruta_s = self.ruta_arch.split(sep="/")
                if ruta_s[-1].split(".")[1] == "xlsx":
                    self.nombre = ruta_s[-1].split(".")[0]
                    del ruta_s[-1]
                    self.dir = "/".join(ruta_s)
                    self.habil_confirm = True
                else:
                    self.ruta_arch = "\nARCHIVO NO ES .xlsx"
                    #raise Exception("Falla en el buscador de archivos.")
        except:
            raise Exception("Falla en el buscador de archivos.")
    
    def volver(self):
        inicio.current = "configur"
        
    def confirmar(self):
        if self.habil_confirm:
            if self.dir and self.nombre:
                self.conf.cfg_custom(self.dir, self.nombre)
                time.sleep(1)
                inicio.add_widget(PesoApp(name = "app"))
                inicio.current = "app"
            else:
                self.ruta_arch = "No hay ruta completa."


# Archivo de configuración
class Confg:
    '''Nombres, directorios y opciones.'''
    def __init__(self) -> None:
        '''Lee/Crea .cfg en dir trabajo.'''
        self.config = configparser.ConfigParser()
        self.RUTA_CFG = os.path.join(RUTA_DIR, NOM_CFG)
    
    def cargar_conf(self):
        '''Carga archivo .cgf en wdir.'''
        print("Cargando configuración...")
        try:
            self.config.read(self.RUTA_CFG)
        except:
            raise Exception("Archivo .cfg no encontrado en directorio de trabajo")
        
        self.dir_xlsx = self.config["RUTAS"]["xlsx"]
        self.nom_xlsx = self.config["NOMBRES"]["xlsx"]
        
        self.dir_bd = self.config["RUTAS"]["bd_sqlite"]
        self.nom_bd = self.config["NOMBRES"]["bd_sqlite"]        
        
        self.sistema = self.config["OPCIONES"]["sistema"]
    
    def cfg_defoult(self):
        '''Guarda .cfg con valores por defecto.'''    
        self.config["NOMBRES"] = NOM_DEFOULT
        # Directorio de ejecución por defercto
        self.config["RUTAS"] = {
            "xlsx": f"{RUTA_DIR}",
            "xlsx_pr": f"{RUTA_DIR}", 
            "bd_sqlite":f"{RUTA_DIR}"
            }
    
        # Guardad sistema (platform)
        self.config["OPCIONES"] = {"sistema":system()}

        with open(self.RUTA_CFG, 'w') as segpeso:
            self.config.write(segpeso)

        print("\nCreado .cgf")

    def cfg_custom(self, dir_xlsx:str, nombre_xlsx:str):
        '''Guarda .cfg con ruta elegida por usuario.'''
        self.config["NOMBRES"] = {
            "xlsx": nombre_xlsx+".xlsx",
            "xlsx_pr": "prueba.xlsx",
            "bd_sqlite":"registro.db"}
        
        self.config["RUTAS"] = {
            "xlsx": f"{dir_xlsx}",
            "xlsx_pr": f"{RUTA_DIR}", 
            "bd_sqlite":f"{RUTA_DIR}"
            }

        # Guardad sistema (platform)
        self.config["OPCIONES"] = {"sistema":system()}

        with open(self.RUTA_CFG, 'w') as segpeso:
            self.config.write(segpeso)

        print("\nCreado .cgf")

# Clases de conexión a bases ###############################
## Conexión con base de datos SQL (ORM)
nombr_bd = "registro.db"
bd = SqliteDatabase(nombr_bd)
class Registro(Model):
    fecha = DateField()
    fecha_regist = CharField()
    peso = FloatField(null = True)
    diametr_sob_ombl_mx = FloatField(null = True)
    diametr_sob_ombl_mn = FloatField(null = True)
    diametr_baj_ombl_mx = FloatField(null = True)
    diametr_baj_ombl_mn = FloatField(null = True)
    class Meta():
        database = bd
        db_table='Medidas'
        primary_key=CompositeKey('fecha', 'fecha_regist')

try:
    bd.connect()
    bd.create_tables([Registro])
    print(f"\n**\nÉxito al conectar con bd:\n {os.path.join(RUTA_DIR,nombr_bd)}\n**\n")
except:
    raise Exception("\nError de Conexión con bd SQL\n")

## Conexión con libro excel
class LibroExcel:
    '''Conexión con archivo excel'''
    def __init__(self, dir_xlsx:str, nom_xlsx:str):
        self.ruta_xlsx= os.path.join(dir_xlsx, nom_xlsx)
        dir_cont = os.listdir(dir_xlsx)
        
        # Algoritmo para crear/leer condicionalmente el xlsx
        if not nom_xlsx in dir_cont:
            print("Archivo excel faltante en directorio elegido.")
            print("Creando... ", "\n\t", self.ruta_xlsx,"\n")
            self.libro = openpyxl.Workbook()
            hoja = self.libro.active
            hoja.title = "Tabla_medidas"
            self.hoja = self.libro["Tabla_medidas"]
            self.hoja.append(("fecha", "peso", "medsomx", "medsomn", "medbomx", "medbomn"))

            self.libro.save(self.ruta_xlsx)

            self.libro = openpyxl.load_workbook(self.ruta_xlsx)
        else:
            print("Cargando excel... : ", "\n\t", self.ruta_xlsx,"\n")
            self.libro = openpyxl.load_workbook(self.ruta_xlsx)

        self.hoja = self.libro["Tabla_medidas"]
    
    def ult_fila(self, hoja):
        for i in range(1, hoja.max_column + 1): 
            cell_obj = hoja.cell(row = hoja.max_row, column = i) 
            print(cell_obj.value, end = " ")

###############################

class Verificar:
    '''Verificación de campos.'''

    def formato(lista_datos:list[str]) -> list[bool]:
        '''Aviso emergente sobre caracteres no válidos y doble coma'''
        l_e = [True for _ in range(len(lista_datos))]
        for i in range(len(lista_datos)):
            if re.search(r'[$%&"\'()¡!¿?#\][/\\]', lista_datos[i]):
                print(f"Valor no válido       \n # EN: {lista_datos[i]} (Verificar.formato)")
                l_e[i] = False 
        for i in range(len(lista_datos)):
            if lista_datos[i].count(".") > 1 or lista_datos[i].count(",") > 1:
                print(lista_datos[i], "tiene +1 dec",  l_e[i])
                l_e[i] = False
        return l_e

    def decim_a_punt(lista_datos:list[str]) -> list[str]:
        '''Corregir marcador decimales.'''
        salida = []
        for dato in  lista_datos:
            if "," in dato:
                dato_modificado = dato.replace(",", ".")
                salida.append(dato_modificado)
            else:
                salida.append(dato)
        return salida
    
    @classmethod
    def verf_nom_xlsx(self, nombre:str) -> str | None:
        '''Avisa de caracteres no válidos y agrega extensión,
        de faltar.'''
        if re.search(r'[$%&"\'¡!,¿?#\][/\\]', nombre):
            print("aviso emergente en verf_nom_xlsx()")
            global nom_inval_pop
            nom_inval_pop = MainApp.dialog_emerg("Nombre inválido", 
                "Introduzca nombre válido")
            nom_inval_pop.open()
            return None
        else:
            if re.search(r'[a-zA-Z0-9].xlsx', nombre):
                nombre = nombre + ".xlsx"
            return nombre
    
    @classmethod
    def cerr_aviso_xlsx(self):
        MainApp.cerrar_dialog(self.nom_inval_pop)


class ConfEmerg(Screen):
    dir_xlsx = StringProperty()
    nombre_xlsx = StringProperty()
    mns_dir = "\n      ruta carpeta ..."
    
    # Guardar prop. 'vacias' para verificar si entraron datos
    _dir_xlsx, _nombre_xlsx = dir_xlsx, nombre_xlsx
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dir_xlsx = self.mns_dir
        self.nombre_xlsx = ""
        
        self.config = Confg()
    
    def configurar(self):
        '''Permite al usuario guardar nombre y directorio para el .xlsx'''

        if self.dir_xlsx == self.mns_dir or self.nombre_xlsx == "":
            global sin_camp_pop
            sin_camp_pop = MainApp.dialog_emerg("Aviso", 
                "Complete los campos faltantes.")
            sin_camp_pop.open()
        else:
            nombre = Verificar.verf_nom_xlsx(self.nombre_xlsx)
            if nombre:
                self.config.cfg_custom(self.dir_xlsx, nombre)
                print(nombre)
                time.sleep(2)
                inicio.add_widget(PesoApp(name = "app"))
                inicio.current = "app"
    
    def buscador_dir(self):
        '''Buscador de dir. dependiente de sistema 
        (debido a problemas con Ubuntu)'''
        
        try:
            if system() == "Windows":
                dir = filedialog.askdirectory()
                if dir: # evita error si se cierra ventana
                    self.dir_xlsx = dir
            elif system() == "Linux":
                dir = filechooser.choose_dir()
                if dir:
                    self.dir_xlsx = dir[0]
        except:
            raise Exception("Falla en el filechooser.")
        
    def volver(self):
        '''Si ya existe configuración vuelve a app, de lo contrario a configuración.'''
        if _config:
            inicio.current = "app"
        else:
            inicio.current = "sinconf"


class MensErr(BoxLayout):
    mens_err = StringProperty()
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.mens_err = kwargs["mens_err"]


class Dialog(BoxLayout):
    mensaje = StringProperty()
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.mensaje = kwargs["mensaje"]


class Crud:
    '''Registrar medidas: al Excel que venía usando,
    y a una base SQL (métodos CRUD).
    \nSQLite\n
    Permite nulos en todos los campos menos fecha.\n 
    La clave primaria es compuesta de la fecha y la
    fecha y hora del registro.
    Usuario no tiene acceso.
    '''
    
    def __init__(self, dir_xlsx:str, nom_xlsx:str):
        self.tb = Registro
        self.lib_excel = LibroExcel(dir_xlsx, nom_xlsx)

    def alta(self, fecha_s:str, datos_v:list):

        # ALTA en SQLite
        try:
            self.tb.create(
                    fecha = fecha_s,
                    fecha_regist = hora(),
                    peso = datos_v[0],
                    diametr_sob_ombl_mx = datos_v[1],
                    diametr_sob_ombl_mn = datos_v[2],
                    diametr_baj_ombl_mx = datos_v[3],
                    diametr_baj_ombl_mn = datos_v[4],
            )
            print("Guardado registro en SQL")
        except:
            raise Exception("ERROR al crear registro")
    
        # ALTA en Excel
        try:
            regis_exc = tuple([fecha_s] + datos_v)
            print(regis_exc)

            self.lib_excel.hoja.append(regis_exc)
            
            self.lib_excel.libro.save(self.lib_excel.ruta_xlsx)
        except:
            raise Exception("Excel: Error de guardado")
        
    def baja(self):
        print(NotImplemented)

    def modificacion(self):
        print(NotImplemented)


# Eventos bontones y declaración de app ###############################
Config.set('graphics', 'width', 700)
Config.set('graphics', 'height', 350)


class PesoApp(Screen):
    '''ROOT'''
    # Esto es un enlace bidireccional (entra al .kv por 
    #  root.fechainput, vuelve como fechainput: fecha.text) *
    fechainput = StringProperty()
    rutaxlsx = StringProperty()

    peso = ObjectProperty(None)
    medsomx = ObjectProperty(None)
    medsomn = ObjectProperty(None)
    medbomx = ObjectProperty(None)
    medbomn = ObjectProperty(None)

    def __init__(self, **kwargs):
        '''root init. Secuencia de configuración inicial y
        conexiones.'''
        super().__init__(**kwargs)
        self.config = Confg()
        print("inicializado class PesoApp")

        try:
            print("\nPESOAPP INIT: .cfg\n")
            self.config.cargar_conf()
        except:
            raise Exception("Error al leer .cfg")
        
        self.sistem = self.config.sistema
        self.fechainput = FECHA_SIS # * el valor defoult
        self.salida_datos = Crud(self.config.dir_xlsx, self.config.nom_xlsx)
        self.rutaxlsx = os.path.join(self.config.dir_xlsx, self.config.nom_xlsx)

    def guardar(self):
        '''Guardar registro de medidas en .db y .xlsx a la vez.'''
        datos = [self.peso.text, self.medsomx.text, self.medsomn.text, 
                 self.medbomx.text, self.medbomn.text]
        print("\nINPUT: ",datos)
        
        ## Verificar campos ##
        err_form = Verificar.formato(datos) 

        # Si las entradas son válidas:
        if False not in err_form:

            # Pasar todos las comas a puntos
            datos_v = Verificar.decim_a_punt(datos)
        
            # introduce nulo (None) para salvar el error de coerción  
            try:
                for i in range(len(datos_v)):
                    if datos_v[i]== '': 
                        datos_v[i] = None
                    else:
                        datos_v[i] = float(datos_v[i])
            except:
                raise Exception("Imposible convertir a float")
            
            # >>> Entrada a Crud.Alta >>>
            try:
                self.salida_datos.alta(self.fechainput, datos_v)
            except:
                raise Exception("Error en alta de registro.")
        
        # Si hay entradas erroneas:
        else:
            errores = []
            print(err_form)
            for i in range(len(err_form)):
                if err_form[i] == False:
                    errores.append(datos[i])
            errores = '\n                 -> '.join(errores)
            PesoApp.adv_emerg(error=f"Valor/es no válido/s:\n                 -> {errores}")

    def comando_cmd(self):
            '''Comando de apertura para el shell del os'''
            print(self.sistem)
            if self.sistem == "Windows":
                rut_compl = os.path.join(self.rutaxlsx)
                print(rut_compl)
                os.system(f'cmd /k start excel.exe {rut_compl}')
            elif self.sistem == "Linux":
                rut_compl = os.path.join(self.rutaxlsx)
                print(rut_compl)
                os.system(f'libreoffice {rut_compl}')            
            else:
                raise Exception("Abrir excel: sin comando válido")

    def abrir_xlsx(self):
        '''Lanzar app excel/equivalente en hilo.'''
        t = threading.Thread(target=self.comando_cmd)
        t.daemon = True
        t.start()

    def configurar(self):
        inicio.current = "configur"

# Declaración de aplicación  #######################
class MainApp(App):
    title = "Seguimiento Peso"
    def build(self):
        # Administrador de pantallas
        global inicio
        inicio = ScreenManager(transition=FadeTransition())
        inicio.add_widget(PrimEjec(name = "sinconf"))
        inicio.add_widget(ConfEmerg(name = "configur"))
        inicio.add_widget(CargarArch(name = "cargar_arch"))

        ## lanzar aviso de config, si no hay .cfg
        if _config:
            print("#####\nintenta iniciar app\n#####")
            inicio.add_widget(PesoApp(name = "app"))
            inicio.current = "app"
        else:
            print("#####\nNo detecta .cfg\n#####")
            inicio.current = "sinconf"

        return inicio 

    # Aviso emergente
    @classmethod
    def dialog_emerg(self, titulo:str, mns:str) -> Popup:
        '''Aviso emergente. Usar .open() para lanzar ventana'''
        print("Aviso emergente: ",mns)
        self.dialog = Popup(title=titulo,
            title_size=20,
            content=Dialog(mensaje=mns),
            size_hint=(None, None),
            size=(300,250))
        
        return self.dialog
        
    @classmethod
    def cerrar_dialog(self): #dialog:Popup
        '''Cierra ventanas emergentes.'''
        if nom_inval_pop:
            nom_inval_pop.dismiss()
        if sin_camp_pop:
            sin_camp_pop.dismiss()
        

if __name__ == '__main__':
    MainApp().run()